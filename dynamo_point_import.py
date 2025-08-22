# -*- coding: utf-8 -*-
"""Dynamo script to create location points in Revit from an Excel file.

This script is intended to run inside a Dynamo Python node. It reads rows from
an Excel worksheet, creates a family instance for each row, and fills the
instance parameters with the provided values.

Inputs
------
excel_path : str
    Absolute path to the Excel file (.xlsx) containing the data.
family_name : str
    Name of the Revit family used for the location point. The family must be
    loaded in the current project and contain the parameters documented in the
    accompanying instructions.

Output
------
list[Element]
    The created Revit family instances.
"""

import clr

# Revit and Dynamo references
clr.AddReference('RevitAPI')
clr.AddReference('RevitServices')
import Autodesk.Revit.DB as DB
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

# Excel handling
import openpyxl

# Retrieve the active document
DOC = DocumentManager.Instance.CurrentDBDocument

# Inputs from Dynamo
excel_path = IN[0]
family_name = IN[1]

# ---------------------------------------------------------------------------
# Helper: read Excel file and return structured data
# ---------------------------------------------------------------------------
def read_excel(path):
    """Return a list of dictionaries representing rows in the Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col_num = header.index('Ponto_Numero')
    col_desc = header.index('Ponto_Descricao')
    col_x = header.index('Coordenada_X')
    col_y = header.index('Coordenada_Y')
    col_z = header.index('Coordenada_Z')

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col_num]:
            continue
        rows.append({
            'num': row[col_num],
            'desc': row[col_desc],
            'x': float(row[col_x]),
            'y': float(row[col_y]),
            'z': float(row[col_z]),
        })
    return rows

# ---------------------------------------------------------------------------
# Helper: fetch the requested family symbol
# ---------------------------------------------------------------------------
def get_symbol(doc, fam_name):
    collector = DB.FilteredElementCollector(doc).OfClass(DB.FamilySymbol)
    for symbol in collector:
        if symbol.Family.Name == fam_name:
            return symbol
    return None

# Read data from Excel
point_data = read_excel(excel_path)

# Acquire the family symbol
symbol = get_symbol(DOC, family_name)
if symbol is None:
    raise ValueError('Family "{}" not found.'.format(family_name))

# Ensure the symbol is active
TransactionManager.Instance.EnsureInTransaction(DOC)
if not symbol.IsActive:
    symbol.Activate()
    DOC.Regenerate()
TransactionManager.Instance.TransactionTaskDone()

# Create family instances and set parameters
created = []
TransactionManager.Instance.EnsureInTransaction(DOC)
for data in point_data:
    loc = DB.XYZ(data['x'], data['y'], data['z'])
    inst = DOC.Create.NewFamilyInstance(
        loc, symbol, DB.Structure.StructuralType.NonStructural)
    inst.LookupParameter('Ponto_Numero').Set(str(data['num']))
    inst.LookupParameter('Ponto_Descricao').Set(str(data['desc']))
    inst.LookupParameter('Ponto_Coordenada_Leste (X)').Set(str(data['x']))
    inst.LookupParameter('Ponto_Coordenada_Norte (Y)').Set(str(data['y']))
    inst.LookupParameter('Ponto_Coordenada_Elevacao (Z)').Set(str(data['z']))
    created.append(inst)
TransactionManager.Instance.TransactionTaskDone()

# Output for Dynamo
OUT = created
